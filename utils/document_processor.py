import os
import PyPDF2
from docx import Document
import openpyxl
from typing import List, Dict, Any
import magic
import uuid
from pathlib import Path
import fitz  # PyMuPDF for better PDF handling
import cv2
import numpy as np
from PIL import Image
import pytesseract
import io

# LangChain text splitters
from langchain_text_splitters import RecursiveCharacterTextSplitter, CharacterTextSplitter, MarkdownHeaderTextSplitter
try:
    # Preferred location in newer LangChain text splitters
    from langchain_text_splitters import SemanticChunker  # type: ignore
except Exception:
    try:
        # Backward compatibility for older LangChain versions
        from langchain_experimental.text_splitter import SemanticChunker  # type: ignore
    except Exception:
        SemanticChunker = None  # Fallback handled at runtime

try:
    # Use Hugging Face embeddings for SemanticChunker
    from langchain_community.embeddings import HuggingFaceEmbeddings  # type: ignore
except Exception:
    HuggingFaceEmbeddings = None  # Will fallback to non-semantic strategies if unavailable

class DocumentProcessor:
    def __init__(self, upload_dir: str = "./uploads"):
        self.upload_dir = upload_dir
        os.makedirs(upload_dir, exist_ok=True)
        
        # Configure Tesseract path for Windows if needed
        try:
            import pytesseract
            # Try to detect Tesseract automatically
            pytesseract.get_tesseract_version()
        except Exception:
            # If automatic detection fails, try common Windows paths
            import platform
            if platform.system() == "Windows":
                common_paths = [
                    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                    r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.getenv('USERNAME', '')),
                ]
                
                for path in common_paths:
                    if os.path.exists(path):
                        pytesseract.pytesseract.tesseract_cmd = path
                        print(f"✓ Found Tesseract at: {path}")
                        break
                else:
                    print("⚠️  Tesseract not found. Please install it or set the path manually.")
                    print("   Download from: https://github.com/UB-Mannheim/tesseract/wiki")
    
    def save_uploaded_file(self, file_content: bytes, original_filename: str) -> Dict[str, Any]:
        """Save uploaded file and return file info"""
        # Generate unique filename
        file_extension = Path(original_filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = os.path.join(self.upload_dir, unique_filename)
        
        # Save file
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # Get file info
        file_size = len(file_content)
        file_type = magic.from_buffer(file_content, mime=True)
        
        return {
            "filename": unique_filename,
            "original_filename": original_filename,
            "file_path": file_path,
            "file_size": file_size,
            "file_type": file_type
        }
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file including images using OCR"""
        try:
            # First try PyMuPDF (fitz) for better text extraction
            try:
                doc = fitz.open(file_path)
                text = ""
                
                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    
                    # Extract text from the page
                    page_text = page.get_text()
                    text += page_text + "\n"
                    
                    # Check if page has images and extract text from them
                    image_list = page.get_images()
                    if image_list:
                        print(f"Page {page_num + 1} contains {len(image_list)} images, extracting text using OCR...")
                        
                        for img_index, img in enumerate(image_list):
                            try:
                                # Get image from page
                                xref = img[0]
                                pix = fitz.Pixmap(doc, xref)
                                
                                if pix.n - pix.alpha < 4:  # GRAY or RGB
                                    # Convert to PIL Image for OCR
                                    img_data = pix.tobytes("png")
                                    pil_image = Image.open(io.BytesIO(img_data))
                                    
                                    # Extract text using Tesseract
                                    img_text = pytesseract.image_to_string(pil_image)
                                    if img_text.strip():
                                        text += f"\n[Image {img_index + 1} Text]: {img_text.strip()}\n"
                                        print(f"✓ Extracted text from image {img_index + 1}: {len(img_text.strip())} characters")
                                
                                pix = None  # Free memory
                                
                            except Exception as img_error:
                                print(f"Warning: Could not extract text from image {img_index + 1}: {str(img_error)}")
                                continue
                
                doc.close()
                
                # If we got text, return it
                if text.strip():
                    return text
                    
            except Exception as fitz_error:
                print(f"PyMuPDF failed, falling back to PyPDF2: {str(fitz_error)}")
            
            # Fallback to PyPDF2 if PyMuPDF fails
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
                
        except Exception as e:
            raise Exception(f"Error extracting text from PDF: {str(e)}")
    
    def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error extracting text from DOCX: {str(e)}")
    
    # def extract_text_from_excel(self, file_path: str) -> str:
    #     """Extract text from Excel file"""
    #     try:
    #         workbook = openpyxl.load_workbook(file_path, data_only=True)
    #         text = ""
    #         for sheet_name in workbook.sheetnames:
    #             sheet = workbook[sheet_name]
    #             text += f"Sheet: {sheet_name}\n"
    #             for row in sheet.iter_rows(values_only=True):
    #                 row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
    #                 if row_text.strip():
    #                     text += row_text + "\n"
    #             text += "\n"
    #         return text
    #     except Exception as e:
    #         raise Exception(f"Error extracting text from Excel: {str(e)}")
    
    def extract_text_from_txt(self, file_path: str) -> str:
        """Extract text from plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except Exception as e:
            raise Exception(f"Error reading text file: {str(e)}")
    
    def extract_text(self, file_path: str, file_type: str) -> str:
        """Extract text from file based on file type"""
        if file_type == "application/pdf":
            return self.extract_text_from_pdf(file_path)
        elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return self.extract_text_from_docx(file_path)
        elif file_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                          "application/vnd.ms-excel"]:
            return self.extract_text_from_excel(file_path)
        elif file_type.startswith("text/"):
            return self.extract_text_from_txt(file_path)
        else:
            raise Exception(f"Unsupported file type: {file_type}")
    
    def chunk_text(self, text: str, chunk_size: int = 1000, overlap: int = 200, 
                   min_chunk_size: int = 100, max_chunk_size: int = 2000) -> List[str]:
        """
        Split text using LangChain's RecursiveCharacterTextSplitter for better semantic chunks
        
        Args:
            text: Text to split
            chunk_size: Target chunk size
            overlap: Overlap between chunks
            min_chunk_size: Minimum acceptable chunk size
            max_chunk_size: Maximum acceptable chunk size
        """
        if len(text) <= chunk_size:
            return [text]
        
        # Clean and normalize text
        text = self._clean_text(text)
        
        # Use LangChain's RecursiveCharacterTextSplitter
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", "! ", "? ", "; ", ", ", " ", ""]
        )
        
        # Split the text
        chunks = text_splitter.split_text(text)
        
        # Filter chunks by size constraints
        chunks = [chunk for chunk in chunks if min_chunk_size <= len(chunk) <= max_chunk_size]
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize text for better chunking"""
        # Remove excessive whitespace
        text = " ".join(text.split())
        
        # Normalize line breaks
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        
        # Remove excessive newlines
        text = "\n".join(line.strip() for line in text.split("\n") if line.strip())
        
        return text
    
    def chunk_text_advanced(self, text: str, strategy: str = "recursive", **kwargs) -> List[str]:
        """
        Advanced text chunking with multiple LangChain strategies
        
        Args:
            text: Text to split
            strategy: Chunking strategy ('recursive', 'semantic', 'fixed', 'markdown')
            **kwargs: Strategy-specific parameters
        """
        if strategy == "recursive":
            return self.chunk_text(text, **kwargs)
        elif strategy == "semantic":
            # Use SemanticChunker if available, otherwise fallback to paragraph-first recursive
            if SemanticChunker is not None and HuggingFaceEmbeddings is not None:
                return self._semantic_chunking_semantic_chunker(text, **kwargs)
            return self._semantic_chunking(text, **kwargs)
        elif strategy == "fixed":
            return self._fixed_size_chunking(text, **kwargs)
        elif strategy == "markdown":
            return self._markdown_chunking(text, **kwargs)
        else:
            raise ValueError(f"Unknown chunking strategy: {strategy}")
    
    def _semantic_chunking(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Semantic chunking using LangChain with paragraph-aware splitting"""
        text = self._clean_text(text)
        
        # Use RecursiveCharacterTextSplitter with paragraph-first approach
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", "! ", "? ", "; ", ", ", " ", ""]
        )
        
        return text_splitter.split_text(text)

    def _semantic_chunking_semantic_chunker(
        self,
        text: str,
        model_name: str = "all-MiniLM-L6-v2",
        breakpoint_threshold_type: str = "percentile",
        breakpoint_threshold_amount: float = 95.0,
    ) -> List[str]:
        """
        Semantic chunking using LangChain's SemanticChunker with Hugging Face embeddings.
        - breakpoint_threshold_type: "percentile" | "standard_deviation" | "interquartile"
        - breakpoint_threshold_amount: higher => fewer, larger chunks; lower => more, smaller chunks
        """
        if SemanticChunker is None or HuggingFaceEmbeddings is None:
            return self._semantic_chunking(text)

        cleaned = self._clean_text(text)
        try:
            embeddings = HuggingFaceEmbeddings(model_name=model_name)
            splitter = SemanticChunker(
                embeddings,
                breakpoint_threshold_type=breakpoint_threshold_type,
                breakpoint_threshold_amount=breakpoint_threshold_amount,
            )
            return splitter.split_text(cleaned)
        except Exception:
            # If anything fails, gracefully fallback
            return self._semantic_chunking(cleaned)
    
    def _fixed_size_chunking(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Fixed-size chunking using LangChain's CharacterTextSplitter"""
        text = self._clean_text(text)
        
        # Use CharacterTextSplitter for fixed-size chunks
        text_splitter = CharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
            separator="\n"
        )
        
        return text_splitter.split_text(text)
    
    def _markdown_chunking(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Markdown-aware chunking using LangChain's MarkdownHeaderTextSplitter"""
        text = self._clean_text(text)
        
        # First split by headers if markdown-like structure exists
        headers_to_split_on = [
            ("#", "Header 1"),
            ("##", "Header 2"),
            ("###", "Header 3"),
            ("####", "Header 4"),
        ]
        
        markdown_splitter = MarkdownHeaderTextSplitter(
            headers_to_split_on=headers_to_split_on
        )
        
        # Split by headers first
        header_splits = markdown_splitter.split_text(text)
        
        # Then apply recursive splitting to each section
        final_chunks = []
        for split in header_splits:
            if len(split.page_content) <= chunk_size:
                final_chunks.append(split.page_content)
            else:
                # Further split large sections
                section_chunks = self.chunk_text(split.page_content, chunk_size, overlap)
                final_chunks.extend(section_chunks)
        
        return final_chunks

    def _analyze_document_structure(self, text: str) -> Dict[str, Any]:
        """Lightweight heuristic analysis to guide intelligent chunking."""
        lines = text.split("\n")
        num_lines = len(lines)
        num_chars = len(text)
        num_headers = sum(1 for ln in lines if ln.strip().startswith(("#", "##", "###", "####", "#####")))
        avg_line_len = (sum(len(ln) for ln in lines) / max(1, num_lines))
        num_paragraphs = len([p for p in text.split("\n\n") if p.strip()])
        has_markdown = num_headers > 0
        is_long = num_chars > 50_000
        is_very_long = num_chars > 200_000
        return {
            "num_lines": num_lines,
            "num_chars": num_chars,
            "num_headers": num_headers,
            "avg_line_len": avg_line_len,
            "num_paragraphs": num_paragraphs,
            "has_markdown": has_markdown,
            "is_long": is_long,
            "is_very_long": is_very_long,
        }

    def chunk_text_intelligent(
        self,
        text: str,
        mode: str = "balanced",
        doc_type: str = "plain",
        hf_model_name: str = "all-MiniLM-L6-v2",
    ) -> List[str]:
        """
        Intelligent chunking that chooses the best strategy based on content and desired trade-off.

        Args:
            mode: "fast" | "balanced" | "quality"
            doc_type: "markdown" | "pdf" | "docx" | "plain"
            hf_model_name: Hugging Face model for SemanticChunker
        """
        cleaned = self._clean_text(text)
        info = self._analyze_document_structure(cleaned)

        # Strategy selection
        if doc_type == "markdown" or info["has_markdown"]:
            # Markdown-aware splitting first
            chunk_size = 1200 if mode == "quality" else 900 if mode == "balanced" else 700
            overlap = 150 if mode != "fast" else 100
            return self._markdown_chunking(cleaned, chunk_size=chunk_size, overlap=overlap)

        # For very long docs and fast mode, favor fixed-size for speed
        if mode == "fast" and (info["is_long"] or info["is_very_long"]):
            return self._fixed_size_chunking(cleaned, chunk_size=800, overlap=80)

        # Prefer semantic chunking when available, with thresholds tuned by mode
        if SemanticChunker is not None and HuggingFaceEmbeddings is not None:
            # Higher breakpoint => fewer chunks (faster ingest), lower => more chunks (higher quality)
            if mode == "quality":
                threshold_amount = 85.0
            elif mode == "fast":
                threshold_amount = 98.0
            else:  # balanced
                threshold_amount = 93.0

            chunks = self._semantic_chunking_semantic_chunker(
                cleaned,
                model_name=hf_model_name,
                breakpoint_threshold_type="percentile",
                breakpoint_threshold_amount=threshold_amount,
            )

            # Safety re-chunking to cap extremely large chunks for embedding efficiency
            max_len = 1600 if mode == "quality" else 1200 if mode == "balanced" else 900
            overlap = 120 if mode != "fast" else 80
            final_chunks: List[str] = []
            for ch in chunks:
                if len(ch) <= max_len:
                    final_chunks.append(ch)
                else:
                    final_chunks.extend(
                        RecursiveCharacterTextSplitter(
                            chunk_size=max_len,
                            chunk_overlap=overlap,
                            length_function=len,
                            separators=["\n\n", "\n", ". ", " "]
                        ).split_text(ch)
                    )
            return final_chunks

        # Fallback to paragraph-first recursive splitting
        chunk_size = 1100 if mode == "quality" else 900 if mode == "balanced" else 750
        overlap = 150 if mode != "fast" else 80
        return self._semantic_chunking(cleaned, chunk_size=chunk_size, overlap=overlap)
